from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "catalogo_grafica_educar.xlsx")

COLUMNS = [
    "Item",
    "Opção",
    "Coleção",
    "Faixa etária / nível",
    "Título",
    "Ilustrador(es) 1",
    "Ilustrador(es) 2",
    "ISBN",
    "Ano de publicação",
    "Número de páginas",
    "Sinopse",
    "Preço unitário",
    "Material de apoio pedagógico",
]

COL_WIDTHS = [8, 8, 22, 22, 44, 30, 30, 20, 18, 16, 60, 16, 34]

COR_HEADER_BG   = "1F4E79"
COR_HEADER_FG   = "FFFFFF"
COR_LINHA_PAR   = "D6E4F0"
COR_LINHA_IMPAR = "FFFFFF"
COR_RESUMO_BG   = "2E75B6"
COR_TOTAL_BG    = "1F4E79"
COR_TOTAL_FG    = "FFFFFF"
COR_SUBHEADER   = "BDD7EE"


def _border():
    s = Side(style="thin", color="B0C4D8")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(cor):
    return PatternFill("solid", fgColor=cor)


def _fmt_header(cell, cor_bg=COR_HEADER_BG, cor_fg=COR_HEADER_FG, size=11):
    cell.fill = _fill(cor_bg)
    cell.font = Font(bold=True, color=cor_fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _fmt_data(cell, row_idx, align="left"):
    cor = COR_LINHA_PAR if row_idx % 2 == 0 else COR_LINHA_IMPAR
    cell.fill = _fill(cor)
    cell.alignment = Alignment(vertical="center", horizontal=align, wrap_text=(align == "left"))
    cell.border = _border()


# ──────────────────────────────────────────────
# ABA 1: CATÁLOGO
# ──────────────────────────────────────────────
def build_catalogo(wb: Workbook, records: list[dict]) -> None:
    ws = wb.active
    ws.title = "Catálogo"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _fmt_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

    CENTER_COLS = {"Item", "Opção", "Ano de publicação", "Número de páginas", "ISBN", "Preço unitário"}

    for row_idx, rec in enumerate(records, start=2):
        ws.row_dimensions[row_idx].height = 18
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = rec.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = "center" if col_name in CENTER_COLS else "left"
            _fmt_data(cell, row_idx, align)

    last_row = len(records) + 1
    last_col = get_column_letter(len(COLUMNS))
    table = Table(displayName="TabelaCatalogo", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.sheet_view.zoomScale = 90


# ──────────────────────────────────────────────
# ABA 2: RESUMO
# ──────────────────────────────────────────────
def build_resumo(wb: Workbook, records: list[dict]) -> None:
    from openpyxl.utils import quote_sheetname
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.zoomScale = 95
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18

    cat = quote_sheetname("Catálogo")

    def title_block(row, col, text, span=2):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _fill(COR_RESUMO_BG)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        ws.row_dimensions[row].height = 28

    def label(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, size=10, color="1F4E79")
        c.fill = _fill(COR_SUBHEADER)
        c.border = _border()
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def val(row, col, formula_or_val, fmt="0", bold=False):
        c = ws.cell(row=row, column=col, value=formula_or_val)
        c.font = Font(bold=bold, size=10)
        c.number_format = fmt
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")

    def total(row, col, lbl, formula):
        for c_idx in range(col, col + 2):
            c = ws.cell(row=row, column=c_idx)
            c.fill = _fill(COR_TOTAL_BG)
            c.font = Font(bold=True, color=COR_TOTAL_FG, size=10)
            c.border = _border()
        ws.cell(row=row, column=col, value=lbl).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=row, column=col + 1, value=formula).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    # Totais gerais
    title_block(1, 1, "TOTAIS GERAIS")
    label(2, 1, "Total de títulos")
    val(2, 2, f"=COUNTA({cat}!E:E)-1", bold=True)
    label(3, 1, "Total de páginas")
    val(3, 2, f"=SUM({cat}!J:J)", bold=True)
    label(4, 1, "Média páginas/título")
    val(4, 2, f"=IFERROR(ROUND(SUM({cat}!J:J)/(COUNTA({cat}!E:E)-1),1),0)", fmt="0.0")

    # Por Tipo (D1)
    tipos = sorted(set(r["Título"].split(" - ")[-1] for r in records))
    title_block(1, 4, "POR TIPO")
    for i, tipo in enumerate(tipos, start=2):
        label(i, 4, tipo)
        val(i, 5, f'=COUNTIF({cat}!E:E,"*{tipo}*")')
        ws.row_dimensions[i].height = 20
    total(len(tipos) + 2, 4, "TOTAL", f"=SUM(E2:E{len(tipos)+1})")

    # Por Faixa etária (A6)
    niveis = sorted(set(r["Faixa etária / nível"] for r in records))
    n_start = 6
    title_block(n_start, 1, "POR FAIXA ETÁRIA")
    for i, nivel in enumerate(niveis, start=n_start + 1):
        label(i, 1, nivel)
        val(i, 2, f'=COUNTIF({cat}!D:D,"{nivel}")')
        ws.row_dimensions[i].height = 20
    total(n_start + len(niveis) + 1, 1, "TOTAL",
          f"=SUM(B{n_start+1}:B{n_start+len(niveis)})")

    # Por Coleção
    colecoes = sorted(set(r["Coleção"] for r in records if r["Coleção"]))
    c_start = n_start + len(niveis) + 3
    title_block(c_start, 1, "POR COLEÇÃO")
    for i, col in enumerate(colecoes, start=c_start + 1):
        label(i, 1, col)
        val(i, 2, f'=COUNTIF({cat}!C:C,"{col}")')
        ws.row_dimensions[i].height = 20
    total(c_start + len(colecoes) + 1, 1, "TOTAL",
          f"=SUM(B{c_start+1}:B{c_start+len(colecoes)})")

    # Por Disciplina (D col)
    disciplinas = sorted(set(r["Título"].split(" - ")[0].split()[-1].upper()
                             if " - " in r["Título"] else "" for r in records) - {""})
    d_start = len(tipos) + 4
    title_block(d_start, 4, "POR ANO")
    anos = sorted(set(r["Ano de publicação"] for r in records if r["Ano de publicação"]))
    for i, ano in enumerate(anos, start=d_start + 1):
        label(i, 4, str(ano))
        val(i, 5, f'=COUNTIF({cat}!I:I,"{ano}")')
        ws.row_dimensions[i].height = 20
    if anos:
        total(d_start + len(anos) + 1, 4, "TOTAL",
              f"=SUM(E{d_start+1}:E{d_start+len(anos)})")


# ──────────────────────────────────────────────
# ABA 3: POR FAIXA ETÁRIA
# ──────────────────────────────────────────────
def build_por_nivel(wb: Workbook, records: list[dict]) -> None:
    ws = wb.create_sheet("Por Faixa Etária")
    ws.sheet_view.zoomScale = 90

    col_defs = [
        ("Faixa etária / nível", 22),
        ("Título", 44),
        ("ISBN", 22),
        ("Ano", 10),
        ("Páginas", 10),
        ("Coleção", 24),
        ("Ilustrador(es) 1", 30),
        ("Sinopse", 60),
    ]

    for col_idx, (name, width) in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        _fmt_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    CENTER = {3, 4, 5}
    prev_nivel = None

    for row_idx, rec in enumerate(records, start=2):
        nivel = rec["Faixa etária / nível"]

        # Separador visual por nível
        if nivel != prev_nivel:
            sep_row = row_idx
            for c in range(1, len(col_defs) + 1):
                cell = ws.cell(row=sep_row, column=c, value=nivel if c == 1 else "")
                cell.fill = _fill("2E75B6")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.border = _border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[sep_row].height = 22
            # Shift subsequent rows
            row_idx += 1
            prev_nivel = nivel

        cor = COR_LINHA_PAR if row_idx % 2 == 0 else COR_LINHA_IMPAR
        values = [
            rec["Faixa etária / nível"],
            rec["Título"],
            rec["ISBN"],
            rec["Ano de publicação"],
            rec["Número de páginas"],
            rec["Coleção"],
            rec["Ilustrador(es) 1"],
            rec["Sinopse"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = _fill(cor)
            cell.border = _border()
            align = "center" if col_idx in CENTER else "left"
            cell.alignment = Alignment(vertical="center", horizontal=align,
                                       wrap_text=(col_idx == 8))
        ws.row_dimensions[row_idx].height = 18


# ──────────────────────────────────────────────
# PONTO DE ENTRADA
# ──────────────────────────────────────────────
def write_excel(records: list[dict]) -> str:
    wb = Workbook()
    print("  Construindo aba Catálogo...")
    build_catalogo(wb, records)
    print("  Construindo aba Resumo...")
    build_resumo(wb, records)
    print("  Construindo aba Por Faixa Etária...")
    build_por_nivel(wb, records)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE
